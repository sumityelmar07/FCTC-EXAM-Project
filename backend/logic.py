# Business logic for FCTC exam automation - INTELLIGENT COLLEGE-AWARE MATCHING
import openpyxl
import os
import re
import logging
from typing import Dict, List, Tuple, Optional
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExamProcessor:
    """
    INTELLIGENT COLLEGE-AWARE MATCHING PIPELINE
    
    ROLL CALL FILE - REQUIRED COLUMNS:
    - "Full name" (MANDATORY for all students)
    - "Roll Number" (MANDATORY for all students)
    - "PRN" (MANDATORY only for VIT students)
    - "Division"
    
    FCTC FILE - REQUIRED COLUMNS:
    - "Full name"
    - "Roll Number"
    - "PRN" (may be missing for non-VIT students)
    - "Score"
    
    INTELLIGENT MATCHING STRATEGY:
    
    Step 1: Data Normalization
    - Trim spaces
    - Uppercase names
    - Clean PRN (remove special chars, decimals)
    - Normalize Roll Number
    - Normalize Division
    
    Step 2: College-Aware Matching
    IF PRN exists AND student belongs to VIT:
        Try PRN match first
    ELSE:
        Skip PRN matching
    
    After PRN step (or if PRN not applicable):
    Level 1: Match using Roll Number + Division (Primary universal identifier)
    Level 2: Exact Full Name match (case-insensitive)
    Level 3: Fuzzy Full Name match using Jaccard similarity (threshold >= 0.8)
             Only within same Division
    
    Step 3: Duplicate Handling
    If multiple FCTC entries for same student: Keep highest score
    
    Step 4: Logging System
    If PRN mismatch but Roll+Div matched: Log event with details
    
    Step 5: Performance
    Use lookup dictionaries for O(n + m) time complexity
    """
    
    def __init__(self):
        self.mismatch_log = []  # Store PRN mismatch events
    
    def _read_excel_with_header_detection(self, file_path: str) -> Tuple[List[List], List[str]]:
        """
        Read Excel file and detect header row automatically with multiple fallback methods
        Returns: (data_rows, header_row)
        """
        print(f"🔍 Attempting to read Excel file: {file_path}")
        
        # Method 1: Try with pandas if available (most robust for problematic files)
        try:
            print("📖 Method 1: Trying pandas (if available)...")
            import pandas as pd
            
            # Try reading with pandas first (handles most Excel issues)
            df = pd.read_excel(file_path, header=None)
            
            if df.empty:
                raise Exception("Excel file is empty")
            
            # Convert DataFrame to list of lists
            all_rows = df.values.tolist()
            
            # Find header row (look for row with most non-empty string values)
            header_row_idx = 0
            max_string_count = 0
            
            for i, row in enumerate(all_rows[:10]):  # Check first 10 rows
                string_count = sum(1 for cell in row if isinstance(cell, str) and str(cell).strip())
                if string_count > max_string_count:
                    max_string_count = string_count
                    header_row_idx = i
            
            headers = [str(cell).strip() if cell is not None and str(cell) != 'nan' else "" for cell in all_rows[header_row_idx]]
            data_rows = all_rows[header_row_idx + 1:]
            
            print(f"✅ Method 1 successful: Found {len(headers)} columns, {len(data_rows)} data rows")
            return data_rows, headers
            
        except ImportError:
            print("📝 pandas not available - using openpyxl methods")
        except Exception as e1:
            print(f"❌ Method 1 failed: {str(e1)}")
        
        # Method 2: Try with openpyxl (data_only=True) - Enhanced for XML issues
        try:
            print("📖 Method 2: Trying openpyxl (data_only=True)...")
            
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            # Get all rows as lists
            all_rows = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):  # Skip completely empty rows
                    all_rows.append(list(row))
            
            workbook.close()
            
            if not all_rows:
                raise Exception("Excel file is empty")
            
            # Find header row
            header_row_idx = 0
            max_string_count = 0
            
            for i, row in enumerate(all_rows[:10]):
                string_count = sum(1 for cell in row if isinstance(cell, str) and cell.strip())
                if string_count > max_string_count:
                    max_string_count = string_count
                    header_row_idx = i
            
            headers = [str(cell).strip() if cell is not None else "" for cell in all_rows[header_row_idx]]
            data_rows = all_rows[header_row_idx + 1:]
            
            print(f"✅ Method 2 successful: Found {len(headers)} columns, {len(data_rows)} data rows")
            return data_rows, headers
            
        except Exception as e2:
            print(f"❌ Method 2 failed: {str(e2)}")
            
            # Method 3: Try with openpyxl (data_only=False)
            try:
                print("📖 Method 3: Trying openpyxl (data_only=False)...")
                
                workbook = openpyxl.load_workbook(file_path, data_only=False)
                sheet = workbook.active
                
                # Get all rows as lists
                all_rows = []
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        all_rows.append(list(row))
                
                workbook.close()
                
                if not all_rows:
                    raise Exception("Excel file is empty")
                
                # Find header row
                header_row_idx = 0
                max_string_count = 0
                
                for i, row in enumerate(all_rows[:10]):
                    string_count = sum(1 for cell in row if isinstance(cell, str) and cell.strip())
                    if string_count > max_string_count:
                        max_string_count = string_count
                        header_row_idx = i
                
                headers = [str(cell).strip() if cell is not None else "" for cell in all_rows[header_row_idx]]
                data_rows = all_rows[header_row_idx + 1:]
                
                print(f"✅ Method 3 successful: Found {len(headers)} columns, {len(data_rows)} data rows")
                return data_rows, headers
                
            except Exception as e3:
                print(f"❌ Method 3 failed: {str(e3)}")
                
                # Method 4: Try reading with read_only mode and ignore XML errors
                try:
                    print("📖 Method 4: Trying openpyxl read-only mode...")
                    
                    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    sheet = workbook.active
                    
                    all_rows = []
                    for row in sheet.iter_rows(values_only=True):
                        if any(cell is not None for cell in row):
                            all_rows.append(list(row))
                    
                    workbook.close()
                    
                    if not all_rows:
                        raise Exception("Excel file is empty")
                    
                    # Find header row
                    header_row_idx = 0
                    max_string_count = 0
                    
                    for i, row in enumerate(all_rows[:10]):
                        string_count = sum(1 for cell in row if isinstance(cell, str) and cell.strip())
                        if string_count > max_string_count:
                            max_string_count = string_count
                            header_row_idx = i
                    
                    headers = [str(cell).strip() if cell is not None else "" for cell in all_rows[header_row_idx]]
                    data_rows = all_rows[header_row_idx + 1:]
                    
                    print(f"✅ Method 4 successful: Found {len(headers)} columns, {len(data_rows)} data rows")
                    return data_rows, headers
                    
                except Exception as e4:
                    print(f"❌ Method 4 failed: {str(e4)}")
                    
                    # Method 5: Try with minimal openpyxl settings and error tolerance
                    try:
                        print("📖 Method 5: Trying openpyxl with error tolerance...")
                        
                        # Try to load with minimal settings
                        import tempfile
                        import shutil
                        
                        # Create a temporary copy
                        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                            shutil.copy2(file_path, tmp_file.name)
                            tmp_path = tmp_file.name
                        
                        try:
                            # Try different openpyxl configurations
                            for keep_vba in [False, True]:
                                for keep_links in [False, True]:
                                    try:
                                        wb = openpyxl.load_workbook(
                                            tmp_path, 
                                            read_only=False, 
                                            keep_vba=keep_vba,
                                            data_only=True,
                                            keep_links=keep_links
                                        )
                                        ws = wb.active
                                        
                                        all_rows = []
                                        for row in ws.iter_rows(values_only=True, max_row=1000):  # Limit rows to avoid memory issues
                                            if any(cell is not None for cell in row):
                                                all_rows.append(list(row))
                                        
                                        wb.close()
                                        
                                        if all_rows:
                                            # Find header row
                                            header_row_idx = 0
                                            max_string_count = 0
                                            
                                            for i, row in enumerate(all_rows[:10]):
                                                string_count = sum(1 for cell in row if isinstance(cell, str) and cell.strip())
                                                if string_count > max_string_count:
                                                    max_string_count = string_count
                                                    header_row_idx = i
                                            
                                            headers = [str(cell).strip() if cell is not None else "" for cell in all_rows[header_row_idx]]
                                            data_rows = all_rows[header_row_idx + 1:]
                                            
                                            print(f"✅ Method 5 successful: Found {len(headers)} columns, {len(data_rows)} data rows")
                                            return data_rows, headers
                                    except:
                                        continue
                            
                            raise Exception("All openpyxl configurations failed")
                            
                        finally:
                            # Clean up temp file
                            try:
                                os.unlink(tmp_path)
                            except:
                                pass
                                
                    except Exception as e5:
                        print(f"❌ Method 5 failed: {str(e5)}")
                        
                        # All methods failed - provide comprehensive error message
                        error_details = f"""
❌ All Excel reading methods failed. This appears to be a severely corrupted Excel file.

🔧 SOLUTIONS TO TRY:

**Option 1 - Fix the Excel file:**
1. Open the file in Microsoft Excel
2. Select all data (Ctrl+A)
3. Copy it (Ctrl+C)  
4. Create a new Excel workbook
5. Paste the data (Ctrl+V)
6. Save as a new .xlsx file
7. Upload the new file

**Option 2 - Convert to CSV:**
1. Open the file in Excel
2. Save As → CSV format
3. Upload the CSV file instead

**Option 3 - Check file integrity:**
1. Ensure the file is not corrupted
2. Try opening in Google Sheets and re-downloading
3. Verify the file is a valid Excel format

**Technical Details:**
• Method 2 (openpyxl data_only=True): {str(e2)}
• Method 3 (openpyxl data_only=False): {str(e3)}
• Method 4 (read_only): {str(e4)}
• Method 5 (error tolerance): {str(e5)}

If none of these solutions work, the original file may be irreparably corrupted.
"""
                        raise Exception(error_details)
    
    def _is_vit_student(self, prn: str, college_name: str = "") -> bool:
        """
        Determine if student belongs to Vishwakarma Institute of Technology
        Based on PRN pattern or college name
        """
        if not prn:
            return False
        
        # VIT PRN patterns (adjust based on actual patterns)
        vit_patterns = [
            r'^VIT',  # Starts with VIT
            r'^\d{10}$',  # 10-digit numeric PRN
            r'^[A-Z]\d{9}$',  # Letter followed by 9 digits
        ]
        
        for pattern in vit_patterns:
            if re.match(pattern, prn):
                return True
        
        # Check college name if available
        if college_name:
            college_upper = college_name.upper()
            if 'VISHWAKARMA' in college_upper or 'VIT' in college_upper:
                return True
        
        return False
    
    def _clean_prn(self, prn_value) -> str:
        """
        Clean and standardize PRN format with robust handling
        Handles: "12413279.0" → "12413279", " ABC-123 " → "ABC123"
        """
        if prn_value is None:
            return ""
        
        # Convert to string and trim
        prn_str = str(prn_value).strip()
        
        # Remove all decimal points and trailing zeros (Excel formatting)
        # "12413279.0" → "12413279", "123.00" → "123"
        if '.' in prn_str:
            try:
                # Try to convert to float then int to remove decimals
                prn_str = str(int(float(prn_str)))
            except (ValueError, TypeError):
                # If conversion fails, just remove decimal part
                prn_str = prn_str.split('.')[0]
        
        # Convert to uppercase
        prn_str = prn_str.upper()
        
        # Remove any non-alphanumeric characters (spaces, dashes, etc.)
        prn_clean = re.sub(r'[^\w]', '', prn_str)
        
        # Filter out invalid values
        if prn_clean in ['', 'NAN', 'NONE', 'NAT', 'NULL', 'N/A', 'NA']:
            return ""
        
        return prn_clean
    
    def _clean_name(self, name_value) -> str:
        """Clean and normalize name for matching"""
        if name_value is None:
            return ""
        
        name_str = str(name_value).strip().upper()
        # Remove extra spaces
        name_str = re.sub(r'\s+', ' ', name_str)
        # Remove special characters but keep spaces
        name_str = re.sub(r'[^A-Z\s]', '', name_str)
        
        # Filter out invalid values
        if name_str in ['', 'NAN', 'NONE', 'NAT', 'NULL', 'N/A', 'NA']:
            return ""
        
        return name_str
    
    def _clean_roll_no(self, roll_no_value) -> str:
        """
        Clean and normalize roll number with robust formatting
        Handles: "01" → "1", "1.0" → "1", " 001 " → "1", " 01" (right-aligned) → "1"
        """
        if roll_no_value is None:
            return ""
        
        # Convert to string and trim (handles right-aligned values)
        roll_str = str(roll_no_value).strip()
        
        # Handle empty or whitespace-only values
        if not roll_str:
            return ""
        
        # Remove all decimal points and trailing zeros (Excel formatting)
        # "1.0" → "1", "123.00" → "123"
        if '.' in roll_str:
            try:
                # Try to convert to float then int to remove decimals
                roll_str = str(int(float(roll_str)))
            except (ValueError, TypeError):
                # If conversion fails, just remove decimal part
                roll_str = roll_str.split('.')[0]
        
        # Convert to uppercase
        roll_str = roll_str.upper()
        
        # Remove any non-alphanumeric characters (spaces, dashes, etc.)
        roll_clean = re.sub(r'[^\w]', '', roll_str)
        
        # Remove leading zeros for consistent matching
        # "001" → "1", "01" → "1", "0A1" → "A1"
        if roll_clean:
            # Check if it's purely numeric
            if roll_clean.isdigit():
                # Convert to int and back to string to remove leading zeros
                roll_clean = str(int(roll_clean))
            else:
                # For alphanumeric, strip leading zeros
                roll_clean = roll_clean.lstrip('0')
                # Handle case where all were zeros
                if not roll_clean:
                    roll_clean = "0"
        
        # Filter out invalid values
        if roll_clean in ['', 'NAN', 'NONE', 'NAT', 'NULL', 'N/A', 'NA']:
            return ""
        
        return roll_clean
    
    def _clean_division(self, division_value) -> str:
        """
        Clean and normalize division with robust formatting
        Handles: "CS-D" → "CSD", "CS D" → "CSD", " cs-d " → "CSD"
        """
        if division_value is None:
            return ""
        
        # Convert to string and trim
        div_str = str(division_value).strip()
        
        # Convert to uppercase
        div_str = div_str.upper()
        
        # Remove ALL spaces, dashes, underscores, and special characters
        # "CS-D" → "CSD", "CS D" → "CSD", "CS_D" → "CSD"
        div_clean = re.sub(r'[\s\-_/\\]', '', div_str)
        
        # Remove any remaining non-alphanumeric characters
        div_clean = re.sub(r'[^\w]', '', div_clean)
        
        # Filter out invalid values
        if div_clean in ['', 'NAN', 'NONE', 'NAT', 'NULL', 'N/A', 'NA']:
            return ""
        
        return div_clean
    
    def _log_prn_mismatch(self, entered_prn: str, matched_prn: str, roll_no: str, 
                          division: str, match_method: str):
        """Log PRN mismatch events for audit"""
        log_entry = {
            'timestamp': datetime.now().isoformat(),
            'entered_prn': entered_prn,
            'matched_prn': matched_prn,
            'roll_number': roll_no,
            'division': division,
            'match_method': match_method
        }
        self.mismatch_log.append(log_entry)
        logger.warning(f"PRN Mismatch: Entered={entered_prn}, Matched={matched_prn}, "
                      f"Roll={roll_no}, Div={division}, Method={match_method}")
    
    def _fuzzy_name_match(self, name1: str, name2: str, threshold: float = 0.8) -> bool:
        """
        Check if two names match using Jaccard similarity
        Returns True if similarity >= threshold
        """
        if not name1 or not name2:
            return False
        
        # Convert to sets of words
        set1 = set(name1.split())
        set2 = set(name2.split())
        
        if not set1 or not set2:
            return False
        
        # Calculate Jaccard similarity
        intersection = len(set1.intersection(set2))
        union = len(set1.union(set2))
        
        if union == 0:
            return False
        
        similarity = intersection / union
        return similarity >= threshold
    
    def _extract_fctc_data(self, data_rows: List[List], headers: List[str]) -> List[Dict]:
        """Extract required fields from FCTC data with intelligent column matching"""
        
        # Create case-insensitive header mapping
        header_map = {}
        for i, header in enumerate(headers):
            if header:
                header_map[header.lower().strip()] = i
        
        # DEBUG: Log all available headers
        logger.info("📋 Available FCTC Headers:")
        for idx, header in enumerate(headers[:15]):  # Show first 15 headers
            logger.info(f"   Column {idx}: '{header}'")
        
        # Define required field mappings (case-insensitive, flexible)
        field_mappings = {
            'full_name': [
                'full name- mandatory for all college students',
                'full name',
                'name',
                'student name'
            ],
            'roll_number': [
                'roll number-mandatory for all college students',
                'roll number',
                'roll no',
                'roll',
                'sr.no',
                'sr no'
            ],
            'prn': [
                'prn - mandatory only for vishwakarma institute of technology students',
                'prn',
                'student prn'
            ],
            'score': [
                'score',
                'total score',
                'marks',
                'total marks'
            ],
            'division_non_vit': [
                'division-mandatory only for non-vishwakarma institute of technology students',
                'division',
                'div',
                'section'
            ],
            'division_vit': [
                'branch-division- mandatory only for vishwakarma institute of technology students',
                'branch-division-mandatory only for vishwakarma institute of technology students',
                'branch-division - mandatory only for vishwakarma institute of technology students',
                'branch-division',
                'branch division'
            ],
            'college_name': [
                'college name-mandatory for all college students ( please select your specific college name carefully and accurately )',
                'college name',
                'college',
                'institution'
            ]
        }
        
        # Find column indices for each field
        column_indices = {}
        for field, possible_names in field_mappings.items():
            for name in possible_names:
                if name in header_map:
                    column_indices[field] = header_map[name]
                    logger.info(f"  ✓ Matched '{field}' to column: '{name}'")
                    break
        
        # Log which division columns were found
        has_vit_division = 'division_vit' in column_indices
        has_non_vit_division = 'division_non_vit' in column_indices
        
        logger.info(f"📋 Division Column Detection:")
        logger.info(f"  VIT Division Column Found: {has_vit_division}")
        if has_vit_division:
            vit_col_idx = column_indices['division_vit']
            logger.info(f"  VIT Division Column Index: {vit_col_idx}")
            logger.info(f"  VIT Division Column Name: '{headers[vit_col_idx]}'")
        logger.info(f"  Non-VIT Division Column Found: {has_non_vit_division}")
        if has_non_vit_division:
            non_vit_col_idx = column_indices['division_non_vit']
            logger.info(f"  Non-VIT Division Column Index: {non_vit_col_idx}")
            logger.info(f"  Non-VIT Division Column Name: '{headers[non_vit_col_idx]}'")
        
        # Check for critical fields
        if 'score' not in column_indices:
            available_cols = list(headers)
            raise Exception(f"❌ FCTC FILE ERROR: Missing required Score column\n"
                          f"Available columns: {available_cols}\n"
                          f"💡 Please ensure your FCTC file has a Score/Marks column.")
        
        if 'full_name' not in column_indices:
            available_cols = list(headers)
            raise Exception(f"❌ FCTC FILE ERROR: Missing required Full Name column\n"
                          f"Available columns: {available_cols}\n"
                          f"💡 Please ensure your FCTC file has a Full Name column.")
        
        if 'roll_number' not in column_indices:
            available_cols = list(headers)
            raise Exception(f"❌ FCTC FILE ERROR: Missing required Roll Number column\n"
                          f"Available columns: {available_cols}\n"
                          f"💡 Please ensure your FCTC file has a Roll Number column.")
        
        # PRN is optional (not all colleges use it)
        if 'prn' not in column_indices:
            logger.info("⚠️ PRN column not found in FCTC file - will use Roll+Division matching")
        
        # Division columns are optional but at least one should exist
        if not has_vit_division and not has_non_vit_division:
            logger.warning("⚠️ No division columns found in FCTC file - matching may be limited")
        
        # Extract data
        extracted_data = []
        student_records = {}  # Track multiple attempts: key = (roll_no, division), value = best record
        
        for row_idx, row in enumerate(data_rows):
            if not any(cell for cell in row):  # Skip empty rows
                continue
                
            # Ensure row has enough columns
            while len(row) < len(headers):
                row.append(None)
            
            # Extract and clean required fields
            full_name_raw = row[column_indices['full_name']] if len(row) > column_indices['full_name'] else None
            full_name_clean = self._clean_name(full_name_raw)
            
            roll_no_raw = row[column_indices['roll_number']] if len(row) > column_indices['roll_number'] else None
            roll_no_clean = self._clean_roll_no(roll_no_raw)
            
            # Skip rows without name or roll number
            if not full_name_clean or not roll_no_clean:
                continue
            
            # Extract PRN (optional)
            prn_raw = row[column_indices['prn']] if 'prn' in column_indices and len(row) > column_indices['prn'] else None
            prn_clean = self._clean_prn(prn_raw)
            
            # Extract College Name (optional) - needed for division column selection
            college_name = row[column_indices['college_name']] if 'college_name' in column_indices and len(row) > column_indices['college_name'] else ""
            
            # DYNAMIC DIVISION COLUMN SELECTION
            # Priority: Try VIT column first (if exists), then Non-VIT column
            # This handles cases where PRN is missing but student is VIT
            division_raw = None
            division_clean = ""
            division_column_used = ""  # Track which column was used
            
            # Check if this is a VIT student
            is_vit = self._is_vit_student(prn_clean, str(college_name))
            
            # PRIORITY 1: If VIT division column exists, try it first
            if has_vit_division:
                division_raw = row[column_indices['division_vit']] if len(row) > column_indices['division_vit'] else None
                division_clean = self._clean_division(division_raw)
                if division_clean:  # Only use if not empty
                    division_column_used = "VIT (Branch-Division)"
            
            # PRIORITY 2: If VIT column was empty or doesn't exist, try Non-VIT column
            if not division_clean and has_non_vit_division:
                division_raw = row[column_indices['division_non_vit']] if len(row) > column_indices['division_non_vit'] else None
                division_clean = self._clean_division(division_raw)
                if division_clean:
                    division_column_used = "Non-VIT (Division)"
            
            # DEBUG: Log first 3 records to show division extraction
            if row_idx < 3:
                logger.info(f"  📝 FCTC Row {row_idx+1}:")
                logger.info(f"     Roll: '{roll_no_raw}' → '{roll_no_clean}'")
                logger.info(f"     Division (RAW): '{division_raw}'")
                logger.info(f"     Division (CLEAN): '{division_clean}'")
                logger.info(f"     Division Column Used: {division_column_used}")
                logger.info(f"     Is VIT: {is_vit}")
                logger.info(f"     College: '{college_name}'")
                logger.info(f"     Generated Key: '{roll_no_clean}_{division_clean}'")
            
            # Extract score and convert to float
            score_raw = row[column_indices['score']] if len(row) > column_indices['score'] else None
            try:
                score = float(score_raw) if score_raw is not None else 0.0
            except (ValueError, TypeError):
                score = 0.0
            
            # Create unique key for duplicate detection
            unique_key = f"{roll_no_clean}_{division_clean}" if division_clean else roll_no_clean
            
            # Handle multiple attempts - keep highest score
            if unique_key in student_records:
                if score > student_records[unique_key]['score']:
                    student_records[unique_key] = {
                        'score': score,
                        'full_name_raw': full_name_raw,
                        'full_name_clean': full_name_clean,
                        'roll_no_raw': roll_no_raw,
                        'roll_no_clean': roll_no_clean,
                        'prn_raw': prn_raw,
                        'prn_clean': prn_clean,
                        'division_raw': division_raw,
                        'division_clean': division_clean,
                        'college_name': college_name
                    }
            else:
                student_records[unique_key] = {
                    'score': score,
                    'full_name_raw': full_name_raw,
                    'full_name_clean': full_name_clean,
                    'roll_no_raw': roll_no_raw,
                    'roll_no_clean': roll_no_clean,
                    'prn_raw': prn_raw,
                    'prn_clean': prn_clean,
                    'division_raw': division_raw,
                    'division_clean': division_clean,
                    'college_name': college_name
                }
        
        # Build final records from best attempts
        for unique_key, data in student_records.items():
            record = {
                'Full_Name_RAW': data['full_name_raw'],
                'Full_Name_CLEAN': data['full_name_clean'],
                'Roll_Number_RAW': data['roll_no_raw'],
                'Roll_Number_CLEAN': data['roll_no_clean'],
                'PRN_RAW': data['prn_raw'],
                'PRN_CLEAN': data['prn_clean'],
                'Division_RAW': data['division_raw'],
                'Division_CLEAN': data['division_clean'],
                'College_Name': data['college_name'],
                'Score': data['score']
            }
            extracted_data.append(record)
        
        if not extracted_data:
            raise Exception("❌ FCTC FILE ERROR: No valid data rows found with required fields")
        
        logger.info(f"✅ Extracted {len(extracted_data)} FCTC records (after duplicate handling)")
        return extracted_data
    
    def _extract_roll_call_data(self, data_rows: List[List], headers: List[str]) -> List[Dict]:
        """Extract roll call data with flexible column matching"""
        
        # Create case-insensitive header mapping
        header_map = {}
        for i, header in enumerate(headers):
            if header:
                header_map[header.lower().strip()] = i
        
        # Define required field mappings (case-insensitive, flexible)
        field_mappings = {
            'full_name': [
                'full name',
                'name',
                'student name',
                'full name (mandatory for all students)'
            ],
            'roll_no': [
                'roll number',
                'roll no',
                'roll',
                'sr.no',
                'sr no',
                'srno',
                'serial no',
                'roll number (mandatory for all students)'
            ],
            'prn': [
                'prn',
                'student prn',
                'prn (mandatory only for vishwakarma institute of technology students)'
            ],
            'division': [
                'division',
                'div',
                'section',
                'class'
            ]
        }
        
        # Find column indices
        column_indices = {}
        for field, possible_names in field_mappings.items():
            for name in possible_names:
                if name in header_map:
                    column_indices[field] = header_map[name]
                    break
        
        # Check for required columns
        missing_cols = []
        for field in ['full_name', 'roll_no', 'division']:
            if field not in column_indices:
                missing_cols.append(field.upper().replace('_', ' '))
        
        if missing_cols:
            available_cols = list(headers)
            variations = {
                'FULL NAME': 'Full Name (or variations: Name, Student Name)',
                'ROLL NO': 'Roll Number (or variations: Roll No, Sr.No, Serial No)',
                'DIVISION': 'Division (or variations: Div, Section, Class)'
            }
            
            missing_details = [f"• '{variations.get(col, col)}'" for col in missing_cols]
            
            raise Exception(f"❌ ROLL CALL FILE ERROR: Missing required columns\n"
                          f"Missing columns:\n" + "\n".join(missing_details) + "\n"
                          f"Available columns in your file:\n" + 
                          "\n".join([f"• '{col}'" for col in available_cols]) + "\n"
                          f"💡 Please ensure your Roll Call file has the required column names.")
        
        # PRN is optional (not all colleges use it)
        if 'prn' not in column_indices:
            logger.info("⚠️ PRN column not found in Roll Call file - will use Roll+Division matching")
        
        # Extract data
        extracted_data = []
        for row in data_rows:
            if not any(cell for cell in row):  # Skip empty rows
                continue
                
            # Ensure row has enough columns
            while len(row) < len(headers):
                row.append(None)
            
            # Extract and clean required fields
            full_name_raw = row[column_indices['full_name']] if len(row) > column_indices['full_name'] else None
            full_name_clean = self._clean_name(full_name_raw)
            
            roll_no_raw = row[column_indices['roll_no']] if len(row) > column_indices['roll_no'] else None
            roll_no_clean = self._clean_roll_no(roll_no_raw)
            
            division_raw = row[column_indices['division']] if len(row) > column_indices['division'] else None
            division_clean = self._clean_division(division_raw)
            
            # Skip rows without required fields
            if not full_name_clean or not roll_no_clean or not division_clean:
                continue
            
            # Extract PRN (optional)
            prn_raw = row[column_indices['prn']] if 'prn' in column_indices and len(row) > column_indices['prn'] else None
            prn_clean = self._clean_prn(prn_raw)
            
            record = {
                'Full_Name_RAW': full_name_raw,
                'Full_Name_CLEAN': full_name_clean,
                'Roll_No_RAW': roll_no_raw,
                'Roll_No_CLEAN': roll_no_clean,
                'PRN_RAW': prn_raw,
                'PRN_CLEAN': prn_clean,
                'Division_RAW': division_raw,
                'Division_CLEAN': division_clean
            }
            
            extracted_data.append(record)
        
        if not extracted_data:
            raise Exception("❌ ROLL CALL FILE ERROR: No valid data rows found with required fields")
        
        logger.info(f"✅ Extracted {len(extracted_data)} Roll Call records")
        return extracted_data
    
    def read_fctc_excel(self, file_path):
        """Read FCTC Excel file and extract ONLY required fields"""
        try:
            data_rows, headers = self._read_excel_with_header_detection(file_path)
            return self._extract_fctc_data(data_rows, headers)
        except Exception as e:
            raise Exception(f"Error reading FCTC Excel file: {str(e)}")
    
    def read_roll_call_excel(self, file_path):
        """Read Roll Call Excel file"""
        try:
            data_rows, headers = self._read_excel_with_header_detection(file_path)
            return self._extract_roll_call_data(data_rows, headers)
        except Exception as e:
            raise Exception(f"Error reading Roll Call Excel file: {str(e)}")
    
    def process_and_generate_reports(self, fctc_file_path, roll_call_file_path, year):
        """
        INTELLIGENT COLLEGE-AWARE MATCHING PIPELINE
        
        Matching Priority:
        1. IF VIT student: Try PRN match first
        2. Roll Number + Division (Primary universal identifier)
        3. Exact Full Name match (case-insensitive)
        4. Fuzzy Full Name match (Jaccard >= 0.8, within same division)
        """
        try:
            # Reset mismatch log
            self.mismatch_log = []
            
            # Read and extract data
            logger.info("📖 Reading FCTC file...")
            fctc_data = self.read_fctc_excel(fctc_file_path)
            logger.info(f"✅ FCTC file processed: {len(fctc_data)} records")
            
            logger.info("📖 Reading Roll Call file...")
            roll_call_data = self.read_roll_call_excel(roll_call_file_path)
            logger.info(f"✅ Roll Call file processed: {len(roll_call_data)} records")
            
            # Create lookup dictionaries for O(n + m) performance
            logger.info("🔍 Creating lookup dictionaries for intelligent matching...")
            
            # Level 1: PRN lookup (only for VIT students)
            fctc_lookup_by_prn = {}
            vit_student_count = 0
            for record in fctc_data:
                prn = record['PRN_CLEAN']
                college = record.get('College_Name', '')
                if prn and self._is_vit_student(prn, college):
                    fctc_lookup_by_prn[prn] = record
                    vit_student_count += 1
            
            # Level 2: Roll Number + Division lookup (PRIMARY universal identifier)
            fctc_lookup_by_roll_div = {}
            for record in fctc_data:
                roll_no = record['Roll_Number_CLEAN']
                division = record['Division_CLEAN']
                if roll_no and division:
                    key = f"{roll_no}_{division}"
                    # If duplicate, keep higher score
                    if key in fctc_lookup_by_roll_div:
                        if record['Score'] > fctc_lookup_by_roll_div[key]['Score']:
                            fctc_lookup_by_roll_div[key] = record
                    else:
                        fctc_lookup_by_roll_div[key] = record
            
            # Level 3: Name lookup grouped by Division (for exact and fuzzy matching)
            fctc_lookup_by_name_div = {}
            for record in fctc_data:
                name = record['Full_Name_CLEAN']
                division = record['Division_CLEAN']
                if name and division:
                    key = f"{name}_{division}"
                    if key not in fctc_lookup_by_name_div:
                        fctc_lookup_by_name_div[key] = []
                    fctc_lookup_by_name_div[key].append(record)
            
            logger.info(f"  ✓ PRN lookup (VIT only): {len(fctc_lookup_by_prn)} entries ({vit_student_count} VIT students)")
            logger.info(f"  ✓ Roll+Division lookup (PRIMARY): {len(fctc_lookup_by_roll_div)} entries")
            logger.info(f"  ✓ Name+Division lookup: {len(fctc_lookup_by_name_div)} entries")
            
            # DEBUG: Show sample keys from each lookup dictionary
            if fctc_lookup_by_prn:
                sample_prn_keys = list(fctc_lookup_by_prn.keys())[:3]
                logger.info(f"  🔍 Sample PRN keys: {sample_prn_keys}")
            
            if fctc_lookup_by_roll_div:
                sample_roll_div_keys = list(fctc_lookup_by_roll_div.keys())[:5]
                logger.info(f"  🔍 Sample Roll+Div keys: {sample_roll_div_keys}")
            
            if fctc_lookup_by_name_div:
                sample_name_div_keys = list(fctc_lookup_by_name_div.keys())[:3]
                logger.info(f"  🔍 Sample Name+Div keys: {sample_name_div_keys}")
            
            # 🔍 DETAILED DEBUG: Check for specific key "1_CSD"
            logger.info("")
            logger.info("=" * 80)
            logger.info("🔍 DETAILED DICTIONARY DEBUG FOR KEY '1_CSD'")
            logger.info("=" * 80)
            logger.info(f"1️⃣ Roll+Division Dictionary Size: {len(fctc_lookup_by_roll_div)}")
            logger.info(f"2️⃣ Does key '1_CSD' exist? {('1_CSD' in fctc_lookup_by_roll_div)}")
            
            # Show all keys that start with "1_"
            keys_starting_with_1 = [k for k in fctc_lookup_by_roll_div.keys() if k.startswith('1_')]
            logger.info(f"3️⃣ All keys starting with '1_': {len(keys_starting_with_1)} found")
            if keys_starting_with_1:
                logger.info(f"   Keys: {sorted(keys_starting_with_1)[:20]}")  # Show first 20
            
            # Show all keys that end with "_CSD"
            keys_ending_with_csd = [k for k in fctc_lookup_by_roll_div.keys() if k.endswith('_CSD')]
            logger.info(f"4️⃣ All keys ending with '_CSD': {len(keys_ending_with_csd)} found")
            if keys_ending_with_csd:
                logger.info(f"   Keys: {sorted(keys_ending_with_csd, key=lambda x: int(x.split('_')[0]) if x.split('_')[0].isdigit() else 999)[:20]}")  # Show first 20, sorted by roll number
            
            # Show sample FCTC records to verify division extraction
            logger.info(f"5️⃣ Sample FCTC records (first 5) to verify division extraction:")
            for idx, record in enumerate(fctc_data[:5]):
                logger.info(f"   Record {idx+1}:")
                logger.info(f"     Roll (RAW): '{record.get('Roll_Number_RAW', 'N/A')}'")
                logger.info(f"     Roll (CLEAN): '{record['Roll_Number_CLEAN']}'")
                logger.info(f"     Division (RAW): '{record.get('Division_RAW', 'N/A')}'")
                logger.info(f"     Division (CLEAN): '{record['Division_CLEAN']}'")
                logger.info(f"     College: '{record.get('College_Name', 'N/A')}'")
                logger.info(f"     Generated Key: '{record['Roll_Number_CLEAN']}_{record['Division_CLEAN']}'")
            
            # Check if there's a record with Roll=1 in FCTC
            fctc_roll_1_all = [rec for rec in fctc_data if rec['Roll_Number_CLEAN'] == '1']
            logger.info(f"6️⃣ Total FCTC records with Roll=1 (normalized): {len(fctc_roll_1_all)}")
            
            # Check if there's a record with Division=CSD in FCTC
            fctc_csd_all = [rec for rec in fctc_data if rec['Division_CLEAN'] == 'CSD']
            logger.info(f"7️⃣ Total FCTC records with Division=CSD (normalized): {len(fctc_csd_all)}")
            if fctc_csd_all:
                # Show what roll numbers exist in CSD
                csd_rolls = sorted([rec['Roll_Number_CLEAN'] for rec in fctc_csd_all], key=lambda x: int(x) if x.isdigit() else 999)
                logger.info(f"   Roll numbers in CSD: {csd_rolls[:30]}")  # Show first 30
            
            # Check for Roll=1 AND Division=CSD
            fctc_1_csd = [rec for rec in fctc_data if rec['Roll_Number_CLEAN'] == '1' and rec['Division_CLEAN'] == 'CSD']
            logger.info(f"8️⃣ FCTC records with Roll=1 AND Division=CSD: {len(fctc_1_csd)}")
            if fctc_1_csd:
                for idx, rec in enumerate(fctc_1_csd):
                    logger.info(f"   Match {idx+1}:")
                    logger.info(f"     Name: {rec['Full_Name_CLEAN']}")
                    logger.info(f"     Roll (RAW): '{rec.get('Roll_Number_RAW', 'N/A')}'")
                    logger.info(f"     Division (RAW): '{rec.get('Division_RAW', 'N/A')}'")
                    logger.info(f"     Score: {rec.get('Score', 'N/A')}")
            else:
                logger.info(f"   ❌ NO MATCH - This is why Roll No 1 CS-D is marked Absent!")
            
            # 🔍 SEARCH FOR SPECIFIC STUDENT BY PRN
            logger.info(f"9️⃣ Searching for student with PRN=12413279:")
            fctc_by_prn_12413279 = [rec for rec in fctc_data if rec['PRN_CLEAN'] == '12413279']
            if fctc_by_prn_12413279:
                for idx, rec in enumerate(fctc_by_prn_12413279):
                    logger.info(f"   FOUND! Entry {idx+1}:")
                    logger.info(f"     Name (RAW): '{rec.get('Full_Name_RAW', 'N/A')}'")
                    logger.info(f"     Name (CLEAN): '{rec['Full_Name_CLEAN']}'")
                    logger.info(f"     Roll (RAW): '{rec.get('Roll_Number_RAW', 'N/A')}'")
                    logger.info(f"     Roll (CLEAN): '{rec['Roll_Number_CLEAN']}'")
                    logger.info(f"     Division (RAW): '{rec.get('Division_RAW', 'N/A')}'")
                    logger.info(f"     Division (CLEAN): '{rec['Division_CLEAN']}'")
                    logger.info(f"     PRN: '{rec['PRN_CLEAN']}'")
                    logger.info(f"     Score: {rec.get('Score', 'N/A')}")
                    logger.info(f"     Generated Key: '{rec['Roll_Number_CLEAN']}_{rec['Division_CLEAN']}'")
            else:
                logger.info(f"   ❌ NOT FOUND in FCTC file!")
            
            # 🔍 SEARCH FOR STUDENT BY NAME (partial match)
            logger.info(f"🔟 Searching for student with name containing 'SWARUP' or 'DIWAN':")
            fctc_by_name = [rec for rec in fctc_data if 'SWARUP' in rec['Full_Name_CLEAN'] or 'DIWAN' in rec['Full_Name_CLEAN']]
            if fctc_by_name:
                logger.info(f"   Found {len(fctc_by_name)} potential matches:")
                for idx, rec in enumerate(fctc_by_name[:5]):  # Show first 5
                    logger.info(f"   Entry {idx+1}:")
                    logger.info(f"     Name: '{rec['Full_Name_CLEAN']}'")
                    logger.info(f"     Roll: '{rec['Roll_Number_CLEAN']}'")
                    logger.info(f"     Division: '{rec['Division_CLEAN']}'")
                    logger.info(f"     PRN: '{rec['PRN_CLEAN']}'")
                    logger.info(f"     Score: {rec.get('Score', 'N/A')}")
            else:
                logger.info(f"   ❌ NO matches found!")
            
            logger.info("=" * 80)
            logger.info("")
            
            # Track division-level statistics for early warning
            division_match_tracker = {}
            
            # Track matching statistics
            match_stats = {
                'prn_matches': 0,
                'roll_div_matches': 0,
                'exact_name_matches': 0,
                'fuzzy_name_matches': 0,
                'no_match': 0
            }
            
            # Group students by division with intelligent matching
            divisions = {}
            
            # DEBUG: Track first few Roll Call entries for debugging
            logger.info("🔍 DEBUG: First 3 Roll Call entries:")
            for idx, roll_record in enumerate(roll_call_data[:3]):
                logger.info(f"  Entry {idx+1}: Roll={roll_record['Roll_No_CLEAN']}, "
                          f"Div={roll_record['Division_CLEAN']}, "
                          f"Name={roll_record['Full_Name_CLEAN'][:30]}, "
                          f"PRN={roll_record['PRN_CLEAN']}")
            
            for roll_record in roll_call_data:
                division = roll_record['Division_CLEAN']
                if not division:
                    division = 'Unknown'
                
                if division not in divisions:
                    divisions[division] = []
                    division_match_tracker[division] = {
                        'total': 0,
                        'matched': 0,
                        'sample_keys_tried': []
                    }
                
                division_match_tracker[division]['total'] += 1
                
                # Extract student info
                prn = roll_record['PRN_CLEAN']
                roll_name = roll_record['Full_Name_CLEAN']
                roll_no = roll_record['Roll_No_CLEAN']
                college = ""  # Not available in roll call
                
                matched_record = None
                match_method = "Not_Found"
                
                # INTELLIGENT MATCHING LOGIC
                
                # Step 1: IF VIT student AND PRN exists, try PRN match first
                if prn and self._is_vit_student(prn, college):
                    if prn in fctc_lookup_by_prn:
                        matched_record = fctc_lookup_by_prn[prn]
                        match_method = "PRN"
                        match_stats['prn_matches'] += 1
                        
                        # Check if Roll+Div also matches (for validation)
                        roll_div_key = f"{roll_no}_{division}"
                        if roll_div_key in fctc_lookup_by_roll_div:
                            matched_by_roll = fctc_lookup_by_roll_div[roll_div_key]
                            # If PRN doesn't match but Roll+Div does, log mismatch
                            if matched_by_roll['PRN_CLEAN'] != prn:
                                self._log_prn_mismatch(
                                    prn, 
                                    matched_by_roll['PRN_CLEAN'],
                                    roll_no,
                                    division,
                                    "Roll+Div"
                                )
                
                # Step 2: Try Roll Number + Division (PRIMARY universal identifier)
                if not matched_record and roll_no and division:
                    key = f"{roll_no}_{division}"
                    
                    # DEBUG: Store sample keys for first student in each division
                    if len(division_match_tracker[division]['sample_keys_tried']) < 3:
                        division_match_tracker[division]['sample_keys_tried'].append(key)
                    
                    # 🔍 SPECIFIC DEBUG FOR ROLL NO 1 IN CS-D
                    if roll_no == "1" and division == "CSD":
                        logger.info("=" * 80)
                        logger.info("🔍 DEBUG: ROLL NO 1 IN DIVISION CS-D")
                        logger.info("=" * 80)
                        logger.info(f"1️⃣ Roll Call Data (RAW):")
                        logger.info(f"   Roll No (RAW): '{roll_record['Roll_No_RAW']}'")
                        logger.info(f"   Division (RAW): '{roll_record['Division_RAW']}'")
                        logger.info(f"   Full Name (RAW): '{roll_record['Full_Name_RAW']}'")
                        logger.info(f"   PRN (RAW): '{roll_record['PRN_RAW']}'")
                        logger.info("")
                        logger.info(f"2️⃣ Roll Call Data (NORMALIZED):")
                        logger.info(f"   Roll No (CLEAN): '{roll_no}'")
                        logger.info(f"   Division (CLEAN): '{division}'")
                        logger.info(f"   Full Name (CLEAN): '{roll_name}'")
                        logger.info(f"   PRN (CLEAN): '{prn}'")
                        logger.info("")
                        logger.info(f"3️⃣ Generated Roll+Division Key:")
                        logger.info(f"   Key: '{key}'")
                        logger.info(f"   Key exists in FCTC lookup? {key in fctc_lookup_by_roll_div}")
                        logger.info("")
                        logger.info(f"4️⃣ All FCTC entries with Roll=1 (normalized):")
                        fctc_roll_1_entries = [rec for rec in fctc_data if rec['Roll_Number_CLEAN'] == "1"]
                        if fctc_roll_1_entries:
                            for idx, entry in enumerate(fctc_roll_1_entries):
                                logger.info(f"   Entry {idx+1}:")
                                logger.info(f"     Roll (RAW): '{entry.get('Roll_Number_RAW', 'N/A')}'")
                                logger.info(f"     Roll (CLEAN): '{entry['Roll_Number_CLEAN']}'")
                                logger.info(f"     Division (RAW): '{entry.get('Division_RAW', 'N/A')}'")
                                logger.info(f"     Division (CLEAN): '{entry['Division_CLEAN']}'")
                                logger.info(f"     Name (RAW): '{entry.get('Full_Name_RAW', 'N/A')}'")
                                logger.info(f"     Name (CLEAN): '{entry['Full_Name_CLEAN']}'")
                                logger.info(f"     PRN (CLEAN): '{entry['PRN_CLEAN']}'")
                                logger.info(f"     Score: {entry.get('Score', 'N/A')}")
                                logger.info(f"     Roll+Div Key: '{entry['Roll_Number_CLEAN']}_{entry['Division_CLEAN']}'")
                        else:
                            logger.info(f"   ❌ NO FCTC entries found with Roll=1 (after normalization)")
                        logger.info("")
                        logger.info(f"5️⃣ All FCTC entries with Division=CSD (normalized):")
                        fctc_csd_entries = [rec for rec in fctc_data if rec['Division_CLEAN'] == "CSD"]
                        logger.info(f"   Found {len(fctc_csd_entries)} entries with Division=CSD")
                        if fctc_csd_entries:
                            logger.info(f"   Sample Roll+Div keys (first 10):")
                            for idx, entry in enumerate(fctc_csd_entries[:10]):
                                logger.info(f"     {entry['Roll_Number_CLEAN']}_{entry['Division_CLEAN']} (Roll RAW: '{entry.get('Roll_Number_RAW', 'N/A')}')")
                        else:
                            logger.info(f"   ❌ NO FCTC entries found with Division=CSD")
                            # Show what divisions exist
                            all_fctc_divs = set(rec['Division_CLEAN'] for rec in fctc_data if rec['Division_CLEAN'])
                            logger.info(f"   Available divisions in FCTC: {sorted(all_fctc_divs)}")
                        logger.info("")
                        logger.info(f"6️⃣ Name Comparison:")
                        name_div_key = f"{roll_name}_{division}"
                        logger.info(f"   Name+Div Key: '{name_div_key}'")
                        logger.info(f"   Key exists in Name lookup? {name_div_key in fctc_lookup_by_name_div}")
                        if name_div_key in fctc_lookup_by_name_div:
                            candidates = fctc_lookup_by_name_div[name_div_key]
                            logger.info(f"   Found {len(candidates)} candidate(s) with exact name match")
                            for idx, cand in enumerate(candidates):
                                logger.info(f"     Candidate {idx+1}: Roll={cand['Roll_Number_CLEAN']}, Score={cand.get('Score', 'N/A')}")
                        else:
                            # Show similar names in same division
                            similar_names = [k for k in fctc_lookup_by_name_div.keys() if k.endswith(f"_{division}")]
                            logger.info(f"   Available names in Division CSD (first 10):")
                            for name_key in similar_names[:10]:
                                logger.info(f"     {name_key.rsplit('_', 1)[0]}")
                        logger.info("=" * 80)
                    
                    if key in fctc_lookup_by_roll_div:
                        matched_record = fctc_lookup_by_roll_div[key]
                        match_method = "Roll+Div"
                        match_stats['roll_div_matches'] += 1
                        division_match_tracker[division]['matched'] += 1
                        
                        # If PRN was provided but didn't match, log it
                        if prn and matched_record['PRN_CLEAN'] and matched_record['PRN_CLEAN'] != prn:
                            self._log_prn_mismatch(
                                prn,
                                matched_record['PRN_CLEAN'],
                                roll_no,
                                division,
                                "Roll+Div"
                            )
                
                # Step 3: Try Exact Full Name match (within same division)
                if not matched_record and roll_name and division:
                    name_div_key = f"{roll_name}_{division}"
                    if name_div_key in fctc_lookup_by_name_div:
                        candidates = fctc_lookup_by_name_div[name_div_key]
                        if len(candidates) == 1:
                            matched_record = candidates[0]
                            match_method = "Exact_Name"
                            match_stats['exact_name_matches'] += 1
                            division_match_tracker[division]['matched'] += 1
                        else:
                            # Multiple candidates with same name - try to match by roll number
                            for candidate in candidates:
                                if candidate['Roll_Number_CLEAN'] == roll_no:
                                    matched_record = candidate
                                    match_method = "Exact_Name"
                                    match_stats['exact_name_matches'] += 1
                                    division_match_tracker[division]['matched'] += 1
                                    break
                            # If still no match, take first candidate
                            if not matched_record:
                                matched_record = candidates[0]
                                match_method = "Exact_Name"
                                match_stats['exact_name_matches'] += 1
                                division_match_tracker[division]['matched'] += 1
                
                # Step 4: Try Fuzzy Full Name match (Jaccard >= 0.8, within same division)
                if not matched_record and roll_name and division:
                    best_similarity = 0.0
                    best_candidate = None
                    
                    # Search within same division only
                    for name_div_key, candidates in fctc_lookup_by_name_div.items():
                        # Extract division from key
                        if not name_div_key.endswith(f"_{division}"):
                            continue
                        
                        # Extract name from key
                        fctc_name = name_div_key.rsplit('_', 1)[0]
                        
                        # Calculate similarity
                        if self._fuzzy_name_match(roll_name, fctc_name, threshold=0.8):
                            # Calculate exact similarity for comparison
                            set1 = set(roll_name.split())
                            set2 = set(fctc_name.split())
                            similarity = len(set1.intersection(set2)) / len(set1.union(set2))
                            
                            if similarity > best_similarity:
                                best_similarity = similarity
                                # If multiple candidates, prefer one with matching roll number
                                for candidate in candidates:
                                    if candidate['Roll_Number_CLEAN'] == roll_no:
                                        best_candidate = candidate
                                        break
                                if not best_candidate:
                                    best_candidate = candidates[0]
                    
                    if best_candidate:
                        matched_record = best_candidate
                        match_method = "Fuzzy_Name"
                        match_stats['fuzzy_name_matches'] += 1
                        division_match_tracker[division]['matched'] += 1
                
                # Determine status and score
                if matched_record:
                    status = "Present"
                    score = matched_record.get('Score', 'N/A')
                else:
                    status = "Absent"
                    score = "N/A"
                    match_stats['no_match'] += 1
                
                # Add to division report
                divisions[division].append({
                    'PRN': roll_record['PRN_RAW'] if roll_record['PRN_RAW'] else 'N/A',
                    'Roll_No': roll_record['Roll_No_RAW'],
                    'Full_Name': roll_record['Full_Name_RAW'],
                    'Division': roll_record['Division_RAW'],
                    'Attendance_Status': status,
                    'Score': score,
                    'Match_Method': match_method
                })
            
            # Log statistics
            logger.info(f"📊 Found {len(divisions)} divisions: {list(divisions.keys())}")
            
            # Check for division-level matching failures
            logger.info("🔍 Division-Level Match Analysis:")
            for div, stats in division_match_tracker.items():
                match_rate = (stats['matched'] / stats['total'] * 100) if stats['total'] > 0 else 0
                logger.info(f"  Division {div}: {stats['matched']}/{stats['total']} matched ({match_rate:.1f}%)")
                
                # WARNING: Division-wide matching failure
                if stats['total'] >= 10 and match_rate < 10:
                    logger.warning(f"  ⚠️ CRITICAL: Division {div} has {match_rate:.1f}% match rate!")
                    logger.warning(f"     Possible formatting mismatch detected.")
                    logger.warning(f"     Sample Roll+Div keys tried: {stats['sample_keys_tried']}")
                    
                    # Show what keys exist in FCTC for this division
                    fctc_keys_for_div = [k for k in fctc_lookup_by_roll_div.keys() if k.endswith(f"_{div}")]
                    if fctc_keys_for_div:
                        logger.warning(f"     FCTC has {len(fctc_keys_for_div)} entries for division {div}")
                        logger.warning(f"     Sample FCTC keys: {fctc_keys_for_div[:5]}")
                    else:
                        logger.warning(f"     FCTC has NO entries with division suffix '_{div}'")
                        # Check if there are similar division names
                        all_divs_in_fctc = set(k.split('_')[-1] for k in fctc_lookup_by_roll_div.keys())
                        logger.warning(f"     Divisions found in FCTC: {sorted(all_divs_in_fctc)}")
            
            logger.info(f"🎯 Matching Statistics:")
            logger.info(f"  ✓ PRN matches (VIT only): {match_stats['prn_matches']}")
            logger.info(f"  ✓ Roll+Division matches (PRIMARY): {match_stats['roll_div_matches']}")
            logger.info(f"  ✓ Exact Name matches: {match_stats['exact_name_matches']}")
            logger.info(f"  ✓ Fuzzy Name matches: {match_stats['fuzzy_name_matches']}")
            logger.info(f"  ✗ No match (Absent): {match_stats['no_match']}")
            
            # Calculate match rate
            total_students = len(roll_call_data)
            total_matched = (match_stats['prn_matches'] + match_stats['roll_div_matches'] + 
                           match_stats['exact_name_matches'] + match_stats['fuzzy_name_matches'])
            match_rate = (total_matched / total_students * 100) if total_students > 0 else 0
            logger.info(f"  📈 Match Rate: {match_rate:.1f}% ({total_matched}/{total_students})")
            
            # Log PRN mismatches if any
            if self.mismatch_log:
                logger.warning(f"⚠️ Found {len(self.mismatch_log)} PRN mismatches (logged for audit)")
            
            # Generate division-wise reports
            division_reports = {}
            for division, students in divisions.items():
                division_reports[division] = {
                    'students': students,
                    'total_students': len(students),
                    'present_count': sum(1 for s in students if s['Attendance_Status'] == 'Present'),
                    'absent_count': sum(1 for s in students if s['Attendance_Status'] == 'Absent')
                }
                logger.info(f"  📋 Division {division}: {len(students)} students "
                          f"({division_reports[division]['present_count']} present, "
                          f"{division_reports[division]['absent_count']} absent)")
            
            # Return results
            result = {
                'success': True,
                'matched_students': total_matched,
                'total_roll_call': len(roll_call_data),
                'total_fctc': len(fctc_data),
                'divisions': list(divisions.keys()),
                'division_count': len(divisions),
                'division_reports': division_reports,
                'match_stats': match_stats,
                'prn_mismatch_log': self.mismatch_log,
                'reports': {
                    'files_created': [f'attendance_report_division_{div}.csv' for div in divisions.keys()],
                    'summary': f"Processed {total_matched} matched students across {len(divisions)} divisions using intelligent college-aware matching"
                }
            }
            
            return result
            
        except Exception as e:
            logger.error(f"Error in intelligent matching processing: {str(e)}")
            raise Exception(f"Error in intelligent matching processing: {str(e)}")
